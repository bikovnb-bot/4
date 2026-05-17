import os
import json
import random
from decimal import Decimal
from datetime import datetime, timedelta, time
import openpyxl

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib import messages
from django.db.models import Q, Count, Avg
from django.db import transaction
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.core.paginator import Paginator
from django.views.decorators.csrf import csrf_protect
from django.views.decorators.cache import never_cache
from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.conf import settings
from django.urls import reverse
from openpyxl.utils import get_column_letter

from .models import ServiceRequest, UsedMaterial, Material, RequestType, RequestHistory, RequestFile, RequestAssignee
from .forms import (
    ServiceRequestForm, UsedMaterialFormSet, ReportForm,
    ImportMaterialsForm, MaterialForm, RequestFileForm, PublicRequestForm
)
from .utils import rate_limit
from .translator import translate_to_russian
from users.models import UserRole
from buildings.models import Building


def generate_new_captcha(lang='ru'):
    operators = ['+', '-', '*']
    op = random.choice(operators)
    if op == '+':
        a = random.randint(1, 20)
        b = random.randint(1, 20)
    elif op == '-':
        a = random.randint(5, 20)
        b = random.randint(1, a)
    else:
        a = random.randint(1, 10)
        b = random.randint(1, 10)
    return {
        'captcha_num1': a,
        'captcha_num2': b,
        'captcha_operator': op,
    }


@login_required
def request_list(request):
    user = request.user
    role = user.profile.role if hasattr(user, 'profile') else None
    qs = ServiceRequest.objects.select_related('building', 'created_by', 'assigned_to')
    if role == UserRole.WORKER:
        qs = qs.filter(Q(assigned_to=user) | Q(assignees__user=user)).distinct()
    elif role not in [UserRole.ADMIN, UserRole.ENGINEER, UserRole.DISPATCHER]:
        qs = qs.filter(created_by=user)

    status = request.GET.get('status')
    executor = request.GET.get('executor')
    priority = request.GET.get('priority')
    search = request.GET.get('search')
    if status:
        qs = qs.filter(status=status)
    if executor:
        qs = qs.filter(assigned_to_id=executor)
    if priority:
        qs = qs.filter(priority=priority)
    if search:
        qs = qs.filter(
            Q(request_number__icontains=search) |
            Q(description__icontains=search)
        )

    executors = User.objects.filter(profile__role=UserRole.WORKER).order_by('username')
    status_choices = ServiceRequest.STATUS_CHOICES
    priority_choices = ServiceRequest.PRIORITY_CHOICES

    paginator = Paginator(qs, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    all_users = User.objects.filter(is_active=True).values(
        'id', 'username', 'first_name', 'last_name'
    ).order_by('username')

    context = {
        'requests': page_obj,
        'status_choices': status_choices,
        'executors': executors,
        'priority_choices': priority_choices,
        'selected_status': status,
        'selected_executor': executor,
        'selected_priority': priority,
        'search_query': search,
        'all_users': list(all_users),
        'user_role': role,
    }
    return render(request, 'requests_app/request_list.html', context)


@login_required
def request_create(request):
    if request.method == 'POST':
        form = ServiceRequestForm(request.user, request.POST)
        if form.is_valid():
            req = form.save(commit=False)
            req.created_by = request.user
            req.status = 'new'
            req.created_at = timezone.now()
            req.save()
            files = request.FILES.getlist('files')
            for f in files:
                RequestFile.objects.create(
                    request=req,
                    file=f,
                    uploaded_by=request.user,
                    description=''
                )
            messages.success(request, f'Заявка {req.request_number} успешно создана.')
            return redirect('requests_app:request_detail', pk=req.pk)
        else:
            messages.error(request, 'Ошибка в форме.')
    else:
        form = ServiceRequestForm(request.user)
    return render(request, 'requests_app/request_form.html', {'form': form, 'title': 'Создание заявки'})


@login_required
def request_edit(request, pk):
    req = get_object_or_404(ServiceRequest, pk=pk)
    user = request.user
    role = user.profile.role if hasattr(user, 'profile') else None
    if req.status == 'closed' and role != UserRole.ADMIN:
        messages.error(request, 'Только администратор может редактировать закрытую заявку.')
        return redirect('requests_app:request_detail', pk=pk)
    if not (role in [UserRole.ADMIN, UserRole.ENGINEER, UserRole.DISPATCHER] or req.created_by == user):
        messages.error(request, 'Нет прав на редактирование.')
        return redirect('requests_app:request_detail', pk=pk)

    if request.method == 'POST':
        form = ServiceRequestForm(request.user, request.POST, instance=req)
        if form.is_valid():
            req = form.save(commit=False)
            req.save()
            delete_files = request.POST.getlist('delete_files')
            for file_id in delete_files:
                try:
                    file_obj = RequestFile.objects.get(id=file_id, request=req)
                    file_obj.file.delete()
                    file_obj.delete()
                except RequestFile.DoesNotExist:
                    pass
            new_files = request.FILES.getlist('files')
            for f in new_files:
                RequestFile.objects.create(
                    request=req,
                    file=f,
                    uploaded_by=request.user,
                    description=''
                )
            messages.success(request, 'Заявка обновлена.')
            return redirect('requests_app:request_detail', pk=req.pk)
    else:
        form = ServiceRequestForm(request.user, instance=req)
    return render(request, 'requests_app/request_form.html', {
        'form': form,
        'title': 'Редактирование заявки',
        'is_edit': True,
        'request_obj': req,
        'files': req.files.all(),
    })


@login_required
def request_detail(request, pk):
    req = get_object_or_404(ServiceRequest.objects.select_related('building', 'created_by', 'assigned_to'), pk=pk)
    user = request.user
    role = user.profile.role if hasattr(user, 'profile') else None
    if role not in [UserRole.ADMIN, UserRole.ENGINEER, UserRole.DISPATCHER]:
        if role == UserRole.WORKER:
            if not (req.assigned_to == user or req.assignees.filter(user=user).exists()):
                messages.error(request, 'Нет доступа.')
                return redirect('requests_app:request_list')
        else:
            if req.created_by != user:
                messages.error(request, 'Нет доступа.')
                return redirect('requests_app:request_list')
    materials_formset = UsedMaterialFormSet(instance=req)
    can_assign = role in [UserRole.ADMIN, UserRole.ENGINEER, UserRole.DISPATCHER] and req.status in ['new', 'in_progress']
    is_executor = (req.assigned_to == user or req.assignees.filter(user=user).exists())
    can_mark_completed = (is_executor or role in [UserRole.ADMIN, UserRole.ENGINEER, UserRole.DISPATCHER]) and req.status == 'in_progress'
    can_suspend = (is_executor or role in [UserRole.ADMIN, UserRole.ENGINEER, UserRole.DISPATCHER]) and req.status == 'in_progress'
    can_resume = role in [UserRole.ADMIN, UserRole.ENGINEER, UserRole.DISPATCHER] and req.status == 'suspended'
    can_close = role in [UserRole.ADMIN, UserRole.ENGINEER, UserRole.DISPATCHER] and req.status == 'completed'
    can_edit = (role in [UserRole.ADMIN, UserRole.ENGINEER, UserRole.DISPATCHER] or req.created_by == user) and not (req.status == 'closed' and role != UserRole.ADMIN)
    history = req.history.all()[:30]
    assignees = req.assignees.all()
    context = {
        'req': req,
        'request_obj': req,
        'materials_formset': materials_formset,
        'can_assign': can_assign,
        'can_mark_completed': can_mark_completed,
        'can_suspend': can_suspend,
        'can_resume': can_resume,
        'can_close': can_close,
        'can_edit': can_edit,
        'attachments': [],
        'history': history,
        'files': req.files.all(),
        'assignees': assignees,
    }
    return render(request, 'requests_app/request_detail.html', context)


@login_required
def request_delete(request, pk):
    req = get_object_or_404(ServiceRequest, pk=pk)
    user = request.user
    role = user.profile.role if hasattr(user, 'profile') else None
    if not (role in [UserRole.ADMIN, UserRole.ENGINEER, UserRole.DISPATCHER] or req.created_by == user):
        messages.error(request, 'Нет прав на удаление.')
        return redirect('requests_app:request_list')
    if request.method == 'POST':
        RequestHistory.objects.create(
            request=req,
            user=request.user,
            action='Заявка удалена',
        )
        req.delete()
        messages.success(request, f'Заявка {req.request_number} удалена.')
        return redirect('requests_app:request_list')
    return render(request, 'requests_app/request_confirm_delete.html', {'req': req})


@login_required
def request_assign(request, pk):
    req = get_object_or_404(ServiceRequest, pk=pk)
    user = request.user
    role = user.profile.role if hasattr(user, 'profile') else None
    if role not in [UserRole.ADMIN, UserRole.ENGINEER, UserRole.DISPATCHER] or req.status not in ['new', 'in_progress']:
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': 'Нет прав или неверный статус'})
        messages.error(request, 'Нет прав для назначения.')
        return redirect('requests_app:request_detail', pk=pk)
    if request.method == 'POST':
        assigned_to_id = request.POST.get('assigned_to')
        if assigned_to_id:
            req.assigned_to_id = assigned_to_id
            if req.status == 'new':
                req.status = 'in_progress'
            req.save()
            executor_name = req.assigned_to.get_full_name() or req.assigned_to.username
            RequestHistory.objects.create(
                request=req,
                user=request.user,
                action=f'Назначен исполнитель: {executor_name}',
            )
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': True, 'message': 'Исполнитель назначен'})
            messages.success(request, 'Исполнитель назначен.')
        else:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'message': 'Выберите исполнителя'})
            messages.error(request, 'Выберите исполнителя.')
        return redirect('requests_app:request_detail', pk=pk)
    return redirect('requests_app:request_detail', pk=pk)


@login_required
def request_mark_completed(request, pk):
    req = get_object_or_404(ServiceRequest, pk=pk)
    user = request.user
    role = user.profile.role if hasattr(user, 'profile') else None
    is_executor = (req.assigned_to == user or req.assignees.filter(user=user).exists())
    if not (is_executor or role in [UserRole.ADMIN, UserRole.ENGINEER, UserRole.DISPATCHER]) or req.status != 'in_progress':
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': 'Нет прав или неверный статус'})
        messages.error(request, 'Нет прав для отметки выполнения.')
        return redirect('requests_app:request_detail', pk=pk)
    if request.method == 'POST':
        req.status = 'completed'
        req.completed_date = timezone.now()
        time_spent = request.POST.get('time_spent')
        if time_spent and time_spent.isdigit():
            req.time_spent = int(time_spent)
        req.save()
        RequestHistory.objects.create(
            request=req,
            user=request.user,
            action='Заявка отмечена как выполненная' + (f' (время: {time_spent} мин)' if time_spent else '')
        )
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': True, 'message': 'Заявка отмечена выполненной'})
        messages.success(request, 'Заявка выполнена.')
        return redirect('requests_app:request_detail', pk=pk)
    return redirect('requests_app:request_detail', pk=pk)


@login_required
def request_suspend(request, pk):
    req = get_object_or_404(ServiceRequest, pk=pk)
    user = request.user
    role = user.profile.role if hasattr(user, 'profile') else None
    is_executor = (req.assigned_to == user or req.assignees.filter(user=user).exists())
    if not (is_executor or role in [UserRole.ADMIN, UserRole.ENGINEER, UserRole.DISPATCHER]) or req.status != 'in_progress':
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': 'Нет прав для приостановки'})
        messages.error(request, 'Нет прав для приостановки.')
        return redirect('requests_app:request_detail', pk=pk)
    if request.method == 'POST':
        reason = request.POST.get('suspension_reason', '').strip()
        if not reason:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'message': 'Укажите причину приостановки'})
            messages.error(request, 'Укажите причину приостановки.')
        else:
            req.status = 'suspended'
            req.suspension_reason = reason
            req.save()
            RequestHistory.objects.create(
                request=req,
                user=request.user,
                action=f'Заявка приостановлена. Причина: {reason}',
            )
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': True, 'message': 'Заявка приостановлена'})
            messages.success(request, f'Заявка №{req.request_number} приостановлена.')
            return redirect('requests_app:request_detail', pk=pk)
    return render(request, 'requests_app/request_suspend.html', {'req': req})


@login_required
def request_resume(request, pk):
    req = get_object_or_404(ServiceRequest, pk=pk)
    user = request.user
    role = user.profile.role if hasattr(user, 'profile') else None
    if role not in [UserRole.ADMIN, UserRole.ENGINEER, UserRole.DISPATCHER]:
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': 'Нет прав для возобновления'})
        messages.error(request, 'Нет прав для возобновления заявки.')
        return redirect('requests_app:request_detail', pk=pk)
    allowed_statuses = ['suspended']
    if role == UserRole.ADMIN:
        allowed_statuses.append('closed')
    if req.status not in allowed_statuses:
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': 'Заявку нельзя возобновить'})
        messages.error(request, 'Заявку нельзя возобновить.')
        return redirect('requests_app:request_detail', pk=pk)
    if request.method == 'POST':
        old_status = req.status
        req.status = 'in_progress'
        req.save()
        RequestHistory.objects.create(
            request=req,
            user=request.user,
            action='Заявка возобновлена' + (' (после закрытия)' if old_status == 'closed' else ''),
        )
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': True, 'message': 'Заявка возобновлена'})
        messages.success(request, f'Заявка №{req.request_number} возобновлена.')
        return redirect('requests_app:request_detail', pk=pk)
    return render(request, 'requests_app/request_resume.html', {'req': req})


@login_required
def request_close(request, pk):
    req = get_object_or_404(ServiceRequest, pk=pk)
    user = request.user
    role = user.profile.role if hasattr(user, 'profile') else None
    if role not in [UserRole.ADMIN, UserRole.ENGINEER, UserRole.DISPATCHER]:
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': 'Нет прав для закрытия'})
        messages.error(request, 'Нет прав для закрытия заявки.')
        return redirect('requests_app:request_detail', pk=pk)
    if req.status != 'completed':
        messages.error(request, 'Закрыть можно только выполненную заявку.')
        return redirect('requests_app:request_detail', pk=pk)
    if request.method == 'POST':
        material_ids = request.POST.getlist('material_id[]')
        quantities = request.POST.getlist('material_quantity[]')
        units = request.POST.getlist('material_unit[]')
        prices = request.POST.getlist('material_price[]')
        materials_used = False
        with transaction.atomic():
            for mat_id, qty_str, unit, price_str in zip(material_ids, quantities, units, prices):
                if not mat_id or not qty_str:
                    continue
                try:
                    qty = Decimal(qty_str.replace(',', '.'))
                    if qty <= 0:
                        continue
                except (ValueError, TypeError):
                    continue
                try:
                    material = Material.objects.get(pk=int(mat_id))
                except (Material.DoesNotExist, ValueError, TypeError):
                    messages.error(request, f'Материал с ID {mat_id} не найден.')
                    return redirect('requests_app:request_close', pk=req.pk)
                if material.quantity_in_stock < qty:
                    messages.error(request, f'Недостаточно материала "{material.name}" на складе (доступно: {material.quantity_in_stock} {material.unit})')
                    return redirect('requests_app:request_close', pk=req.pk)
                UsedMaterial.objects.create(
                    request=req,
                    material=material,
                    name=material.name,
                    quantity=qty,
                    unit=unit,
                    price_per_unit=Decimal(price_str) if price_str else Decimal(0)
                )
                material.quantity_in_stock -= qty
                material.save()
                materials_used = True
        req.status = 'closed'
        req.save()
        RequestHistory.objects.create(
            request=req,
            user=request.user,
            action='Заявка закрыта' + (' (с материалами)' if materials_used else ' (без материалов)')
        )
        messages.success(request, f'Заявка #{req.request_number} закрыта.')
        return redirect('requests_app:request_detail', pk=req.pk)
    materials_qs = Material.objects.all().values('id', 'name', 'unit', 'default_price')
    materials_json = list(materials_qs)
    context = {
        'request_obj': req,
        'req': req,
        'materials': materials_qs,
        'materials_json': materials_json,
    }
    return render(request, 'requests_app/request_close.html', context)


@login_required
def request_dashboard(request):
    from users.models import UserRole
    
    user = request.user
    role = user.profile.role if hasattr(user, 'profile') else None

    if role not in [UserRole.ADMIN, UserRole.ENGINEER, UserRole.DISPATCHER]:
        messages.error(request, 'У вас нет доступа к дашборду.')
        return redirect('requests_app:request_list')

    qs = ServiceRequest.objects.all()

    total_requests = qs.count()
    completed_closed = qs.filter(status__in=['completed', 'closed']).count()
    in_progress_count = qs.filter(status='in_progress').count()
    overdue_count = qs.filter(
        planned_date__lt=timezone.now().date(),
        status__in=['new', 'in_progress', 'suspended']
    ).count()

    # Статусы
    status_counts = qs.values('status').annotate(total=Count('id'))
    status_labels = []
    status_data = []
    status_display = dict(ServiceRequest.STATUS_CHOICES)
    for item in status_counts:
        status_labels.append(status_display.get(item['status'], item['status']))
        status_data.append(item['total'])

    # Топ-5 типов
    type_counts = qs.values('request_type__name').annotate(total=Count('id')).order_by('-total')[:5]
    type_labels = [item['request_type__name'] or 'Без типа' for item in type_counts]
    type_data = [item['total'] for item in type_counts]

    # Приоритеты
    priority_counts = qs.values('priority').annotate(count=Count('id'))
    priority_labels = []
    priority_data = []
    priority_display = dict(ServiceRequest.PRIORITY_CHOICES)
    for p in priority_counts:
        priority_labels.append(priority_display.get(p['priority'], p['priority']))
        priority_data.append(p['count'])

    # Динамика по месяцам
    today = timezone.now()
    start_date = today - timedelta(days=365)
    requests_in_period = qs.filter(created_at__date__gte=start_date, created_at__date__lte=today)
    from collections import defaultdict
    monthly_dict = defaultdict(int)
    for req in requests_in_period:
        month_key = req.created_at.strftime('%Y-%m')
        monthly_dict[month_key] += 1
    month_labels = sorted(monthly_dict.keys())
    month_data = [monthly_dict[m] for m in month_labels]

    # Среднее время выполнения
    completed_reqs = qs.filter(status__in=['completed', 'closed'], completed_date__isnull=False)
    avg_days = 0
    if completed_reqs.exists():
        total_seconds = sum((req.completed_date - req.created_at).total_seconds() for req in completed_reqs)
        avg_days = total_seconds / len(completed_reqs) / 86400

    # Динамика среднего времени по месяцам
    monthly_avg = []
    for month in month_labels:
        year = int(month[:4])
        month_num = int(month[5:7])
        month_reqs = qs.filter(
            created_at__year=year,
            created_at__month=month_num,
            completed_date__isnull=False,
            status__in=['completed', 'closed']
        )
        if month_reqs.exists():
            total_sec = sum((r.completed_date - r.created_at).total_seconds() for r in month_reqs)
            avg = total_sec / len(month_reqs) / 86400
            monthly_avg.append(round(avg, 1))
        else:
            monthly_avg.append(0)

    # Все рабочие (роль WORKER)
    workers = User.objects.filter(profile__role=UserRole.WORKER).select_related('profile')
    worker_stats = []
    total_completed_closed = completed_closed
    for worker in workers:
        completed_count = ServiceRequest.objects.filter(
            assigned_to=worker,
            status__in=['completed', 'closed']
        ).count()
        total_assigned = ServiceRequest.objects.filter(assigned_to=worker).count()
        percent = (completed_count / total_completed_closed * 100) if total_completed_closed > 0 else 0
        worker_stats.append({
            'name': worker.get_full_name() or worker.username,
            'completed': completed_count,
            'total_assigned': total_assigned,
            'percent': round(percent, 1)
        })
    worker_stats.sort(key=lambda x: x['completed'], reverse=True)

    # Топ-5 исполнителей (по назначенным заявкам)
    top_executors_raw = (
        qs.filter(assigned_to__isnull=False)
        .values('assigned_to__username', 'assigned_to__first_name', 'assigned_to__last_name')
        .annotate(count=Count('id'))
        .order_by('-count')[:5]
    )
    top_executors = []
    for te in top_executors_raw:
        name = f"{te['assigned_to__first_name']} {te['assigned_to__last_name']}".strip()
        if not name:
            name = te['assigned_to__username']
        top_executors.append({'name': name, 'count': te['count']})

    context = {
        'total_requests': total_requests,
        'completed_closed': completed_closed,
        'in_progress_count': in_progress_count,
        'overdue_count': overdue_count,
        'status_labels': status_labels,
        'status_data': status_data,
        'type_labels': type_labels,
        'type_data': type_data,
        'priority_labels': priority_labels,
        'priority_data': priority_data,
        'month_labels': month_labels,
        'month_data': month_data,
        'avg_completion_time': round(avg_days, 1),
        'monthly_avg': monthly_avg,
        'worker_stats': worker_stats,
        'top_executors': top_executors,
    }
    return render(request, 'requests_app/dashboard.html', context)


@login_required
def export_requests_excel(request):
    user = request.user
    role = user.profile.role if hasattr(user, 'profile') else None
    qs = ServiceRequest.objects.select_related('building', 'created_by', 'assigned_to')
    if role == UserRole.WORKER:
        qs = qs.filter(Q(assigned_to=user) | Q(assignees__user=user)).distinct()
    elif role not in [UserRole.ADMIN, UserRole.ENGINEER, UserRole.DISPATCHER]:
        qs = qs.filter(created_by=user)
    status = request.GET.get('status')
    executor = request.GET.get('executor')
    priority = request.GET.get('priority')
    search = request.GET.get('search')
    if status:
        qs = qs.filter(status=status)
    if executor:
        qs = qs.filter(assigned_to_id=executor)
    if priority:
        qs = qs.filter(priority=priority)
    if search:
        qs = qs.filter(
            Q(request_number__icontains=search) |
            Q(description__icontains=search)
        )
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Заявки"
    headers = [
        '№ заявки', 'Здание', 'Помещение', 'Тип', 'Описание', 'Приоритет', 'Статус',
        'Создатель', 'Ответственный', 'Плановая дата', 'Дата выполнения', 'Дата создания'
    ]
    ws.append(headers)
    for req in qs:
        ws.append([
            req.request_number,
            str(req.building),
            req.room_number,
            req.request_type.name if req.request_type else '',
            req.description[:100],
            req.get_priority_display(),
            req.get_status_display(),
            req.created_by.get_full_name() or req.created_by.username,
            req.assigned_to.get_full_name() if req.assigned_to else '',
            req.planned_date.strftime('%d.%m.%Y') if req.planned_date else '',
            req.completed_date.strftime('%d.%m.%Y %H:%M') if req.completed_date else '',
            req.created_at.strftime('%d.%m.%Y %H:%M') if req.created_at else '',
        ])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="requests_export.xlsx"'
    wb.save(response)
    return response


@login_required
def custom_report(request):
    user = request.user
    role = user.profile.role if hasattr(user, 'profile') else None
    if role not in [UserRole.ADMIN, UserRole.ENGINEER, UserRole.DISPATCHER]:
        messages.error(request, 'У вас нет доступа к отчётам.')
        return redirect('requests_app:request_list')
    form = ReportForm(request.GET or None)
    if role == UserRole.WORKER:
        qs = ServiceRequest.objects.filter(
            Q(assigned_to=user) | Q(assignees__user=user)
        ).distinct()
    else:
        qs = ServiceRequest.objects.all()
    if request.GET and form.is_valid():
        if form.cleaned_data.get('status'):
            qs = qs.filter(status=form.cleaned_data['status'])
        if form.cleaned_data.get('priority'):
            qs = qs.filter(priority=form.cleaned_data['priority'])
        if form.cleaned_data.get('building'):
            qs = qs.filter(building=form.cleaned_data['building'])
        if form.cleaned_data.get('request_type'):
            qs = qs.filter(request_type=form.cleaned_data['request_type'])
        if form.cleaned_data.get('assigned_to'):
            qs = qs.filter(assigned_to=form.cleaned_data['assigned_to'])
        if form.cleaned_data.get('created_by'):
            qs = qs.filter(created_by=form.cleaned_data['created_by'])
        if form.cleaned_data.get('room_number'):
            qs = qs.filter(room_number__icontains=form.cleaned_data['room_number'])
        if form.cleaned_data.get('date_from'):
            qs = qs.filter(created_at__date__gte=form.cleaned_data['date_from'])
        if form.cleaned_data.get('date_to'):
            qs = qs.filter(created_at__date__lte=form.cleaned_data['date_to'])
    columns = form.cleaned_data.get('columns') if form.is_valid() else []
    if not columns:
        columns = ['request_number', 'building', 'priority', 'status', 'created_by', 'assigned_to', 'created_at']
    field_map = {
        'request_number': '№ заявки',
        'building': 'Здание',
        'room_number': 'Помещение',
        'request_type': 'Тип',
        'description': 'Описание',
        'priority': 'Приоритет',
        'status': 'Статус',
        'created_by': 'Создатель',
        'assigned_to': 'Ответственный',
        'planned_date': 'Плановая дата',
        'completed_date': 'Дата выполнения',
        'created_at': 'Дата создания',
        'comment': 'Комментарий',
    }
    if request.GET.get('export') == '1':
        import openpyxl
        from openpyxl.utils import get_column_letter
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Отчёт по заявкам"
        for col_idx, col in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=field_map.get(col, col))
            cell.font = openpyxl.styles.Font(bold=True)
        for row_idx, req in enumerate(qs.select_related('building', 'created_by', 'assigned_to', 'request_type'), 2):
            for col_idx, col in enumerate(columns, 1):
                value = ''
                if col == 'request_number':
                    value = req.request_number
                elif col == 'building':
                    value = str(req.building)
                elif col == 'room_number':
                    value = req.room_number or ''
                elif col == 'request_type':
                    value = req.request_type.name if req.request_type else ''
                elif col == 'description':
                    value = req.description[:200] if req.description else ''
                elif col == 'priority':
                    value = req.get_priority_display()
                elif col == 'status':
                    value = req.get_status_display()
                elif col == 'created_by':
                    value = req.created_by.get_full_name() if req.created_by else ''
                elif col == 'assigned_to':
                    value = req.assigned_to.get_full_name() if req.assigned_to else ''
                elif col == 'planned_date':
                    value = req.planned_date.strftime('%d.%m.%Y') if req.planned_date else ''
                elif col == 'completed_date':
                    value = req.completed_date.strftime('%d.%m.%Y %H:%M') if req.completed_date else ''
                elif col == 'created_at':
                    value = req.created_at.strftime('%d.%m.%Y %H:%M') if req.created_at else ''
                elif col == 'comment':
                    value = req.comment[:200] if req.comment else ''
                ws.cell(row=row_idx, column=col_idx, value=value)
        for col_idx in range(1, len(columns) + 1):
            max_length = 0
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="custom_report.xlsx"'
        wb.save(response)
        return response
    data = []
    for req in qs.select_related('building', 'created_by', 'assigned_to', 'request_type'):
        row = {}
        for col in columns:
            if col == 'request_number':
                row[col] = req.request_number
            elif col == 'building':
                row[col] = str(req.building)
            elif col == 'room_number':
                row[col] = req.room_number or ''
            elif col == 'request_type':
                row[col] = req.request_type.name if req.request_type else ''
            elif col == 'description':
                row[col] = req.description[:100] if req.description else ''
            elif col == 'priority':
                row[col] = req.get_priority_display()
            elif col == 'status':
                row[col] = req.get_status_display()
            elif col == 'created_by':
                row[col] = req.created_by.get_full_name() if req.created_by else ''
            elif col == 'assigned_to':
                row[col] = req.assigned_to.get_full_name() if req.assigned_to else ''
            elif col == 'planned_date':
                row[col] = req.planned_date.strftime('%d.%m.%Y') if req.planned_date else ''
            elif col == 'completed_date':
                row[col] = req.completed_date.strftime('%d.%m.%Y %H:%M') if req.completed_date else ''
            elif col == 'created_at':
                row[col] = req.created_at.strftime('%d.%m.%Y %H:%M') if req.created_at else ''
            elif col == 'comment':
                row[col] = req.comment[:200] if req.comment else ''
            else:
                row[col] = ''
        data.append(row)
    context = {
        'form': form,
        'data': data,
        'columns': columns,
        'column_labels': field_map,
        'field_map': field_map,
    }
    return render(request, 'requests_app/custom_report.html', context)


@login_required
def import_materials(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        form = ImportMaterialsForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['excel_file']
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            created = 0
            updated = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                name = row[0] if row[0] else None
                unit = row[1] if len(row) > 1 else None
                default_price = row[2] if len(row) > 2 and row[2] is not None else 0
                quantity_in_stock = row[3] if len(row) > 3 and row[3] is not None else 0
                if name and unit:
                    try:
                        default_price = float(str(default_price).replace(',', '.'))
                    except (ValueError, TypeError):
                        default_price = 0.0
                    try:
                        quantity_in_stock = float(str(quantity_in_stock).replace(',', '.'))
                    except (ValueError, TypeError):
                        quantity_in_stock = 0.0
                    material, is_created = Material.objects.update_or_create(
                        name=name,
                        defaults={
                            'unit': unit,
                            'default_price': default_price,
                            'quantity_in_stock': quantity_in_stock
                        }
                    )
                    if is_created:
                        created += 1
                    else:
                        updated += 1
            messages.success(request, f'Импортировано: добавлено {created}, обновлено {updated}.')
            return redirect('requests_app:material_stock')
        else:
            messages.error(request, 'Ошибка в форме. Проверьте файл.')
    else:
        form = ImportMaterialsForm()
    return render(request, 'requests_app/import_materials.html', {'form': form})


@login_required
def download_materials_template(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Материалы"
    headers = ['name', 'unit', 'default_price', 'quantity_in_stock']
    ws.append(headers)
    ws.append(['Краска', 'л', 350.00, 100])
    ws.append(['Лампа светодиодная', 'шт', 450.00, 50])
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_len:
                    max_len = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 30)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="materials_import_template.xlsx"'
    wb.save(response)
    return response


@login_required
def material_stock(request):
    search_query = request.GET.get('search', '').strip()
    materials_qs = Material.objects.all().order_by('name')
    if search_query:
        materials_qs = materials_qs.filter(name__icontains=search_query)
    paginator = Paginator(materials_qs, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'requests_app/material_stock.html', {
        'materials': page_obj,
        'search': search_query,
    })


@login_required
def material_stock_export(request):
    materials = Material.objects.all().values('name', 'unit', 'quantity_in_stock', 'default_price')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Склад материалов"
    headers = ['Наименование', 'Единица измерения', 'Количество на складе', 'Цена за единицу']
    ws.append(headers)
    for m in materials:
        ws.append([
            m['name'],
            m['unit'],
            float(m['quantity_in_stock']),
            float(m['default_price'])
        ])
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[col_letter].width = adjusted_width
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="material_stock_export.xlsx"'
    wb.save(response)
    return response


@login_required
def material_add(request):
    role = request.user.profile.role if hasattr(request.user, 'profile') else None
    if role not in [UserRole.ADMIN, UserRole.ENGINEER, UserRole.DISPATCHER]:
        messages.error(request, 'Нет прав для добавления материалов.')
        return redirect('requests_app:material_stock')
    if request.method == 'POST':
        form = MaterialForm(request.POST)
        if form.is_valid():
            material = form.save()
            messages.success(request, f'Материал "{material.name}" добавлен.')
            return redirect('requests_app:material_stock')
    else:
        form = MaterialForm()
    return render(request, 'requests_app/material_form.html', {'form': form, 'title': 'Добавить материал'})


@login_required
def material_edit(request, pk):
    material = get_object_or_404(Material, pk=pk)
    user = request.user
    role = user.profile.role if hasattr(user, 'profile') else None
    if role not in [UserRole.ADMIN, UserRole.ENGINEER, UserRole.DISPATCHER]:
        messages.error(request, 'Нет прав для редактирования материала.')
        return redirect('requests_app:material_stock')
    if request.method == 'POST':
        form = MaterialForm(request.POST, instance=material)
        if form.is_valid():
            form.save()
            messages.success(request, f'Материал "{material.name}" успешно обновлён.')
            return redirect('requests_app:material_stock')
        else:
            messages.error(request, 'Ошибка в форме.')
    else:
        form = MaterialForm(instance=material)
    return render(request, 'requests_app/material_form.html', {'form': form, 'title': 'Редактировать материал'})


@login_required
def material_delete(request, pk):
    material = get_object_or_404(Material, pk=pk)
    user = request.user
    role = user.profile.role if hasattr(user, 'profile') else None
    if role not in [UserRole.ADMIN, UserRole.ENGINEER, UserRole.DISPATCHER]:
        messages.error(request, 'Нет прав для удаления материала.')
        return redirect('requests_app:material_stock')
    if request.method == 'POST':
        name = material.name
        material.delete()
        messages.success(request, f'Материал "{name}" удалён.')
        return redirect('requests_app:material_stock')
    return render(request, 'requests_app/material_confirm_delete.html', {'material': material})


@login_required
def material_delete_ajax(request, pk):
    material = get_object_or_404(Material, pk=pk)
    user = request.user
    role = user.profile.role if hasattr(user, 'profile') else None
    if role not in [UserRole.ADMIN, UserRole.ENGINEER, UserRole.DISPATCHER]:
        return JsonResponse({'success': False, 'error': 'Нет прав'}, status=403)
    if request.method == 'POST':
        material.delete()
        return JsonResponse({'success': True})
    return JsonResponse({'success': False, 'error': 'Метод не разрешён'}, status=405)


@login_required
def request_add_assignee(request, pk):
    req = get_object_or_404(ServiceRequest, pk=pk)
    user = request.user
    role = user.profile.role if hasattr(user, 'profile') else None
    if role not in [UserRole.ADMIN, UserRole.ENGINEER, UserRole.DISPATCHER]:
        messages.error(request, 'Нет прав для назначения исполнителей.')
        return redirect('requests_app:request_detail', pk=pk)
    if request.method == 'POST':
        user_id = request.POST.get('user_id')
        if user_id:
            try:
                assignee = User.objects.get(pk=user_id)
                obj, created = RequestAssignee.objects.get_or_create(request=req, user=assignee)
                if created:
                    RequestHistory.objects.create(
                        request=req,
                        user=request.user,
                        action=f'Добавлен исполнитель: {assignee.get_full_name() or assignee.username}'
                    )
                    messages.success(request, f'Исполнитель {assignee.get_full_name() or assignee.username} добавлен.')
                else:
                    messages.warning(request, 'Этот исполнитель уже назначен.')
            except User.DoesNotExist:
                messages.error(request, 'Пользователь не найден.')
        else:
            messages.error(request, 'Выберите пользователя.')
        return redirect('requests_app:request_detail', pk=pk)
    assigned_user_ids = req.assignees.values_list('user_id', flat=True)
    available_users = User.objects.filter(is_active=True).exclude(id__in=assigned_user_ids).exclude(id=req.assigned_to_id).order_by('username')
    return render(request, 'requests_app/add_assignee.html', {'request_obj': req, 'users': available_users})


@login_required
def request_remove_assignee(request, pk, user_id):
    req = get_object_or_404(ServiceRequest, pk=pk)
    user = request.user
    role = user.profile.role if hasattr(user, 'profile') else None
    if role not in [UserRole.ADMIN, UserRole.ENGINEER, UserRole.DISPATCHER]:
        messages.error(request, 'Нет прав для удаления исполнителей.')
        return redirect('requests_app:request_detail', pk=pk)
    assignee = get_object_or_404(RequestAssignee, request=req, user_id=user_id)
    assignee_name = assignee.user.get_full_name() or assignee.user.username
    assignee.delete()
    RequestHistory.objects.create(
        request=req,
        user=request.user,
        action=f'Удалён исполнитель: {assignee_name}'
    )
    messages.success(request, f'Исполнитель {assignee_name} удалён.')
    return redirect('requests_app:request_detail', pk=pk)


# ---------- Публичная форма для неавторизованных пользователей ----------
@csrf_protect
@never_cache
@rate_limit()
def public_request_create(request):
    lang = request.GET.get('lang', 'ru')
    if lang not in ['ru', 'en']:
        lang = 'ru'

    if request.method == 'POST':
        form = PublicRequestForm(request.POST, request.FILES, lang=lang)
        if form.is_valid():
            # Валидация файлов
            files = request.FILES.getlist('files')
            max_files = 5
            max_size = 5 * 1024 * 1024
            allowed_types = ['image/jpeg', 'image/png', 'image/gif', 'application/pdf']
            allowed_extensions = ['.jpg', '.jpeg', '.png', '.gif', '.pdf']

            if len(files) > max_files:
                messages.error(request, f'Можно прикрепить не более {max_files} файлов.')
                post_data = request.POST.copy()
                post_data.update(generate_new_captcha(lang))
                form = PublicRequestForm(post_data, request.FILES, lang=lang)
                context = {'form': form, 'lang': lang, 'hide_navbar': True}
                return render(request, f'requests_app/public_request_form_{lang}.html', context)

            for f in files:
                if f.size > max_size:
                    messages.error(request, f'Файл "{f.name}" превышает 5 МБ.')
                    post_data = request.POST.copy()
                    post_data.update(generate_new_captcha(lang))
                    form = PublicRequestForm(post_data, request.FILES, lang=lang)
                    context = {'form': form, 'lang': lang, 'hide_navbar': True}
                    return render(request, f'requests_app/public_request_form_{lang}.html', context)

                ext = os.path.splitext(f.name)[1].lower()
                if ext not in allowed_extensions:
                    messages.error(request, f'Файл "{f.name}" имеет недопустимое расширение. Разрешены: {", ".join(allowed_extensions)}')
                    post_data = request.POST.copy()
                    post_data.update(generate_new_captcha(lang))
                    form = PublicRequestForm(post_data, request.FILES, lang=lang)
                    context = {'form': form, 'lang': lang, 'hide_navbar': True}
                    return render(request, f'requests_app/public_request_form_{lang}.html', context)

                try:
                    import magic
                    f.seek(0)
                    mime = magic.from_buffer(f.read(1024), mime=True)
                    f.seek(0)
                    if mime not in allowed_types:
                        messages.error(request, f'Файл "{f.name}" имеет недопустимый тип ({mime}).')
                        post_data = request.POST.copy()
                        post_data.update(generate_new_captcha(lang))
                        form = PublicRequestForm(post_data, request.FILES, lang=lang)
                        context = {'form': form, 'lang': lang, 'hide_navbar': True}
                        return render(request, f'requests_app/public_request_form_{lang}.html', context)
                except ImportError:
                    pass

            sr = form.save(commit=False)
            sr.created_by = None
            sr.assigned_to = None
            sr.status = 'new'
            sr.priority = 'low'
            sr.ip_address = request.META.get('REMOTE_ADDR')
            sr.description = translate_to_russian(sr.description)

            contact_info = f"Контактное лицо: {sr.contact_name or 'не указано'}, Телефон: {sr.contact_phone or 'не указан'}"
            if sr.comment:
                sr.comment = f"{sr.comment}\n{contact_info}"
            else:
                sr.comment = contact_info

            sr.save()

            for f in files:
                RequestFile.objects.create(
                    request=sr,
                    file=f,
                    uploaded_by=None,
                    description='Загружено из публичной формы'
                )

            if settings.DEFAULT_FROM_EMAIL and hasattr(settings, 'ADMINS') and settings.ADMINS:
                try:
                    subject = f"New public request #{sr.request_number}"
                    html_message = render_to_string('requests_app/email_new_public_request.html', {'request': sr})
                    send_mail(
                        subject,
                        f"Request #{sr.request_number} from {sr.contact_name or 'Anonymous'}",
                        settings.DEFAULT_FROM_EMAIL,
                        [email for name, email in settings.ADMINS],
                        fail_silently=True,
                        html_message=html_message
                    )
                except Exception:
                    pass

            # Определяем, подана ли заявка в нерабочее время
            now = timezone.localtime(timezone.now())
            off_hours = False
            if now.weekday() >= 5:  # суббота (5) или воскресенье (6)
                off_hours = True
            else:
                work_start = time(9, 30)
                work_end = time(18, 0)
                current_time = now.time()
                if current_time < work_start or current_time >= work_end:
                    off_hours = True

            off_param = '&off_hours=1' if off_hours else ''
            return redirect(f'{reverse("requests_app:public_request_success")}?lang={lang}{off_param}')
        else:
            # Форма не валидна – перегенерируем капчу
            post_data = request.POST.copy()
            post_data.update(generate_new_captcha(lang))
            form = PublicRequestForm(post_data, request.FILES, lang=lang)
            context = {'form': form, 'lang': lang, 'hide_navbar': True}
            return render(request, f'requests_app/public_request_form_{lang}.html', context)
    else:
        form = PublicRequestForm(lang=lang)
        context = {'form': form, 'lang': lang, 'hide_navbar': True}
        return render(request, f'requests_app/public_request_form_{lang}.html', context)


def public_request_success(request):
    lang = request.GET.get('lang', 'ru')
    if lang not in ['ru', 'en']:
        lang = 'ru'
    off_hours = request.GET.get('off_hours') == '1'
    return render(request, f'requests_app/public_request_success_{lang}.html', {'hide_navbar': True, 'off_hours': off_hours})
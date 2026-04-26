from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from decimal import Decimal
from django.urls import reverse, reverse_lazy
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import ListView, DetailView, UpdateView, DeleteView
from django.http import JsonResponse, HttpResponse
from django.template.loader import render_to_string
from django.contrib.auth.models import User
from django.db.models import Sum
from collections import defaultdict
from datetime import datetime, timedelta, date
from dateutil.relativedelta import relativedelta
from calendar import monthrange
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.utils import timezone

from .models import Meter, Reading, ZoneReading, TariffComponent, ResourceType, MeterDocument
from .forms import ReadingForm, MeterForm, ReadingEditForm, MeterDocumentForm, ImportReadingsForm
from .utils import (
    can_view_all_meters, can_edit_all_meters, can_assign_owner,
    can_view_meter, can_edit_meter, can_delete_meter,
    can_edit_reading, can_delete_reading,
    get_avg_consumption, is_anomaly,
    log_action,  # добавлено
)
from users.decorators import manager_required, viewer_required


# ------------------------------------------------------------
# Добавление показаний (обычная страница)
# ------------------------------------------------------------
@login_required
@viewer_required
def add_reading(request):
    initial = {}
    if request.method == 'GET':
        meter_id = request.GET.get('meter')
        date_str = request.GET.get('date')
        if meter_id:
            try:
                meter = Meter.objects.get(pk=int(meter_id))
                if not can_view_meter(request.user, meter):
                    messages.error(request, "У вас нет доступа к этому счётчику.")
                    return redirect('energy:meter_list')
                initial['meter'] = meter.id
            except Meter.DoesNotExist:
                pass
        if date_str:
            initial['date'] = date_str

    if request.method == 'POST':
        form = ReadingForm(request.user, request.POST)
        if form.is_valid():
            reading = form.save()
            # Проверка на аномальное потребление
            avg = get_avg_consumption(reading.meter)
            consumption = reading.total_consumption()
            if is_anomaly(consumption, avg):
                messages.warning(
                    request,
                    f'⚠️ Внимание! Потребление за {reading.date} ({consumption:.2f} {reading.meter.resource_type.unit}) '
                    f'значительно превышает среднее ({avg:.2f} {reading.meter.resource_type.unit})!'
                )
            # Логирование
            log_action(
                user=request.user,
                action='CREATE',
                model_name='Reading',
                object_id=reading.pk,
                details=f"Показание для {reading.meter.serial_number} от {reading.date}: {reading.total_consumption()}",
                request=request
            )
            messages.success(request, "Показания успешно добавлены.")
            return redirect('energy:meter_detail', pk=reading.meter.id)
        else:
            return render(request, 'energy/add_reading.html', {'form': form})
    else:
        form = ReadingForm(request.user, initial=initial)
    return render(request, 'energy/add_reading.html', {'form': form})


# ------------------------------------------------------------
# Добавление показаний через модальное окно (AJAX)
# ------------------------------------------------------------
@login_required
@viewer_required
def add_reading_modal(request):
    if request.method == 'POST':
        form = ReadingForm(request.user, request.POST)
        if form.is_valid():
            reading = form.save()
            avg = get_avg_consumption(reading.meter)
            consumption = reading.total_consumption()
            if is_anomaly(consumption, avg):
                # Для AJAX сообщение не выводим, но можно вернуть предупреждение
                pass
            # Логирование
            log_action(
                user=request.user,
                action='CREATE',
                model_name='Reading',
                object_id=reading.pk,
                details=f"Показание для {reading.meter.serial_number} от {reading.date}: {reading.total_consumption()}",
                request=request
            )
            return JsonResponse({'status': 'success', 'message': 'Показания добавлены'})
        else:
            return JsonResponse({'status': 'error', 'errors': form.errors}, status=400)
    else:
        meter_id = request.GET.get('meter_id')
        initial = {}
        if meter_id:
            try:
                meter = Meter.objects.get(pk=int(meter_id))
                if not can_view_meter(request.user, meter):
                    return JsonResponse({'status': 'error', 'message': 'Нет доступа'}, status=403)
                initial['meter'] = meter.id
            except Meter.DoesNotExist:
                pass
        form = ReadingForm(request.user, initial=initial)
        html = render_to_string('energy/add_reading_modal_form.html', {'form': form}, request=request)
        return JsonResponse({'html': html})


# ------------------------------------------------------------
# Редактирование показаний
# ------------------------------------------------------------
@login_required
@manager_required
def edit_reading(request, pk):
    reading = get_object_or_404(Reading, pk=pk)
    if not can_edit_reading(request.user, reading):
        messages.error(request, "У вас нет прав на редактирование этого показания.")
        return redirect('energy:meter_list')
    if request.method == 'POST':
        form = ReadingEditForm(request.user, reading, request.POST)
        if form.is_valid():
            form.save()
            avg = get_avg_consumption(reading.meter)
            consumption = reading.total_consumption()
            if is_anomaly(consumption, avg):
                messages.warning(
                    request,
                    f'⚠️ Внимание! Потребление за {reading.date} ({consumption:.2f} {reading.meter.resource_type.unit}) '
                    f'значительно превышает среднее ({avg:.2f} {reading.meter.resource_type.unit})!'
                )
            # Логирование
            log_action(
                user=request.user,
                action='EDIT',
                model_name='Reading',
                object_id=reading.pk,
                details=f"Показание для {reading.meter.serial_number} от {reading.date}: {reading.total_consumption()}",
                request=request
            )
            messages.success(request, 'Показания обновлены.')
            next_url = request.POST.get('next', reverse('energy:meter_detail', args=[reading.meter.id]))
            return redirect(next_url)
    else:
        form = ReadingEditForm(request.user, reading)
    return render(request, 'energy/reading_edit.html', {'form': form, 'reading': reading})


# ------------------------------------------------------------
# Удаление показаний
# ------------------------------------------------------------
@login_required
@manager_required
def delete_reading(request, pk):
    reading = get_object_or_404(Reading, pk=pk)
    if not can_delete_reading(request.user, reading):
        messages.error(request, "У вас нет прав на удаление этого показания.")
        return redirect('energy:meter_list')
    if request.method == 'POST':
        next_url = request.POST.get('next', reverse('energy:meter_detail', args=[reading.meter.id]))
        serial = reading.meter.serial_number
        reading.delete()
        log_action(
            user=request.user,
            action='DELETE',
            model_name='Reading',
            object_id=pk,
            details=f"Удалено показание для {serial} от {reading.date}",
            request=request
        )
        messages.success(request, f'Показание счётчика "{serial}" удалено.')
        return redirect(next_url)
    return render(request, 'energy/reading_confirm_delete.html', {'reading': reading})


# ------------------------------------------------------------
# Добавление счётчика
# ------------------------------------------------------------
@login_required
@manager_required
def add_meter(request):
    if request.method == 'POST':
        form = MeterForm(request.user, request.POST)
        if form.is_valid():
            meter = form.save()
            log_action(
                user=request.user,
                action='CREATE',
                model_name='Meter',
                object_id=meter.pk,
                details=f"Добавлен счётчик: {meter.serial_number}, тип: {meter.resource_type.name}",
                request=request
            )
            messages.success(request, f'Счётчик "{meter.serial_number}" успешно добавлен.')
            return redirect('energy:meter_list')
        else:
            messages.error(request, 'Пожалуйста, исправьте ошибки в форме.')
    else:
        form = MeterForm(request.user)
    return render(request, 'energy/add_meter.html', {'form': form})


# ------------------------------------------------------------
# Загрузка документов для счётчика
# ------------------------------------------------------------
@login_required
@manager_required
def upload_document(request, meter_pk):
    meter = get_object_or_404(Meter, pk=meter_pk)
    if not can_edit_meter(request.user, meter):
        messages.error(request, "У вас нет прав на загрузку документов для этого счётчика.")
        return redirect('energy:meter_detail', pk=meter.pk)
    if request.method == 'POST':
        form = MeterDocumentForm(request.POST, request.FILES)
        if form.is_valid():
            doc = form.save(commit=False)
            doc.meter = meter
            doc.save()
            log_action(
                user=request.user,
                action='CREATE',
                model_name='MeterDocument',
                object_id=doc.pk,
                details=f"Загружен документ для счётчика {meter.serial_number}: {doc.get_file_name()}",
                request=request
            )
            messages.success(request, f'Документ "{doc.get_file_name()}" загружен.')
            return redirect('energy:meter_detail', pk=meter.pk)
    else:
        form = MeterDocumentForm()
    return render(request, 'energy/upload_document.html', {'form': form, 'meter': meter})


# ------------------------------------------------------------
# Удаление документа
# ------------------------------------------------------------
@login_required
@manager_required
def delete_document(request, pk):
    doc = get_object_or_404(MeterDocument, pk=pk)
    if not can_edit_meter(request.user, doc.meter):
        messages.error(request, "У вас нет прав на удаление этого документа.")
        return redirect('energy:meter_detail', pk=doc.meter.pk)
    meter_pk = doc.meter.pk
    file_name = doc.get_file_name()
    doc.delete()
    log_action(
        user=request.user,
        action='DELETE',
        model_name='MeterDocument',
        object_id=pk,
        details=f"Удалён документ {file_name} для счётчика {doc.meter.serial_number}",
        request=request
    )
    messages.success(request, 'Документ удалён.')
    return redirect('energy:meter_detail', pk=meter_pk)


# ------------------------------------------------------------
# Список счётчиков (ListView)
# ------------------------------------------------------------
class MeterListView(LoginRequiredMixin, ListView):
    model = Meter
    template_name = 'energy/meter_list.html'
    context_object_name = 'meters'

    def get_queryset(self):
        if can_view_all_meters(self.request.user):
            return Meter.objects.all()
        return Meter.objects.all()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['can_edit'] = can_edit_all_meters(self.request.user)
        context['can_delete'] = can_delete_meter(self.request.user, None)
        return context


# ------------------------------------------------------------
# Детальная информация о счётчике (DetailView) с интерактивным графиком и аномалиями
# ------------------------------------------------------------
class MeterDetailView(LoginRequiredMixin, DetailView):
    model = Meter
    template_name = 'energy/meter_detail.html'
    context_object_name = 'meter'

    def get_queryset(self):
        if can_view_all_meters(self.request.user):
            return Meter.objects.all()
        return Meter.objects.all()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        meter = self.object
        context['can_edit'] = can_edit_meter(self.request.user, meter)
        context['can_delete'] = can_delete_meter(self.request.user, meter)
        context['user'] = self.request.user

        # Логирование просмотра
        log_action(
            user=self.request.user,
            action='VIEW',
            model_name='Meter',
            object_id=meter.pk,
            details=f"Просмотр счётчика {meter.serial_number}",
            request=self.request
        )

        # --- Интерактивный график: параметры периода и типа ---
        months_count = int(self.request.GET.get('months', 12))
        months_count = min(months_count, 36)
        chart_type = self.request.GET.get('chart_type', 'line')
        context['selected_months'] = months_count
        context['chart_type'] = chart_type

        today = date.today()
        start_date = today.replace(day=1) - timedelta(days=1)
        for _ in range(months_count - 1):
            start_date = start_date.replace(day=1) - timedelta(days=1)
        start_date = start_date.replace(day=1)

        readings = meter.reading_set.filter(date__gte=start_date).order_by('date')

        if meter.is_multi_tariff:
            zones = TariffComponent.objects.filter(resource_type=meter.resource_type, is_multi_tariff_zone=True).order_by('name')
            zone_names = [zone.name for zone in zones]
            monthly_data = defaultdict(lambda: defaultdict(float))
            for reading in readings:
                month_key = reading.date.strftime('%Y-%m')
                for zr in reading.zone_readings.all():
                    if zr.consumption:
                        monthly_data[month_key][zr.tariff_component.name] += float(zr.consumption)
            months = sorted(monthly_data.keys())
            zone_datasets = []
            for zone in zone_names:
                data = [monthly_data[month].get(zone, 0) for month in months]
                zone_datasets.append({
                    'label': zone,
                    'data': data,
                    'backgroundColor': f'rgba({hash(zone) % 255}, {(hash(zone) * 2) % 255}, {(hash(zone) * 3) % 255}, 0.6)',
                    'borderColor': f'rgba({hash(zone) % 255}, {(hash(zone) * 2) % 255}, {(hash(zone) * 3) % 255}, 1)',
                    'borderWidth': 1,
                    'stack': 'zones',
                })
            context['month_labels'] = months
            context['zone_datasets'] = zone_datasets
            context['is_multi_tariff'] = True
        else:
            monthly_consumption = defaultdict(float)
            for reading in readings:
                if reading.consumption:
                    month_key = reading.date.strftime('%Y-%m')
                    monthly_consumption[month_key] += float(reading.consumption)
            months = sorted(monthly_consumption.keys())
            consumptions = [monthly_consumption[month] for month in months]
            context['month_labels'] = months
            context['consumptions'] = consumptions
            context['is_multi_tariff'] = False

        # --- Последние 5 показаний с флагом аномалии ---
        last_readings = meter.reading_set.order_by('-date')[:5]
        last_readings_with_anomaly = []
        if last_readings:
            avg = get_avg_consumption(meter)
            for reading in last_readings:
                consumption = reading.total_consumption()
                anomaly = is_anomaly(consumption, avg)
                last_readings_with_anomaly.append({'reading': reading, 'anomaly': anomaly})
        context['last_readings_with_anomaly'] = last_readings_with_anomaly

        return context


# ------------------------------------------------------------
# Редактирование счётчика (UpdateView)
# ------------------------------------------------------------
class MeterUpdateView(LoginRequiredMixin, UpdateView):
    model = Meter
    form_class = MeterForm
    template_name = 'energy/meter_edit.html'
    success_url = reverse_lazy('energy:meter_list')

    def get_queryset(self):
        if can_edit_all_meters(self.request.user):
            return Meter.objects.all()
        return Meter.objects.all()

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        response = super().form_valid(form)
        log_action(
            user=self.request.user,
            action='EDIT',
            model_name='Meter',
            object_id=self.object.pk,
            details=f"Редактирование счётчика {self.object.serial_number}",
            request=self.request
        )
        return response


# ------------------------------------------------------------
# Удаление счётчика (DeleteView)
# ------------------------------------------------------------
class MeterDeleteView(LoginRequiredMixin, DeleteView):
    model = Meter
    template_name = 'energy/meter_confirm_delete.html'
    success_url = reverse_lazy('energy:meter_list')

    def get_queryset(self):
        if can_edit_all_meters(self.request.user):
            return Meter.objects.all()
        return Meter.objects.all()

    def delete(self, request, *args, **kwargs):
        obj = self.get_object()
        serial = obj.serial_number
        log_action(
            user=request.user,
            action='DELETE',
            model_name='Meter',
            object_id=obj.pk,
            details=f"Удаление счётчика {serial}",
            request=request
        )
        return super().delete(request, *args, **kwargs)


# ------------------------------------------------------------
# Отчёт по потреблению (сравнение с прошлым годом)
# ------------------------------------------------------------
@login_required
@viewer_required
def consumption_report(request):
    period_type = request.GET.get('period', 'month')
    year = int(request.GET.get('year', datetime.now().year))
    month = int(request.GET.get('month', datetime.now().month))
    quarter = int(request.GET.get('quarter', (datetime.now().month - 1) // 3 + 1))
    group_id = request.GET.get('group', 'all')

    month_names_ru = {
        1: 'Январь', 2: 'Февраль', 3: 'Март', 4: 'Апрель',
        5: 'Май', 6: 'Июнь', 7: 'Июль', 8: 'Август',
        9: 'Сентябрь', 10: 'Октябрь', 11: 'Ноябрь', 12: 'Декабрь'
    }

    if period_type == 'month':
        start_date = date(year, month, 1)
        end_date = start_date.replace(day=monthrange(year, month)[1])
        period_label = f"{month_names_ru[month]} {year}"
        prev_start_date = date(year-1, month, 1)
        prev_end_date = prev_start_date.replace(day=monthrange(year-1, month)[1])
    elif period_type == 'quarter':
        start_month = (quarter - 1) * 3 + 1
        start_date = date(year, start_month, 1)
        end_date = start_date + relativedelta(months=3) - timedelta(days=1)
        period_label = f"{quarter} квартал {year}"
        prev_start_date = date(year-1, start_month, 1)
        prev_end_date = prev_start_date + relativedelta(months=3) - timedelta(days=1)
    else:
        start_date = date(year, 1, 1)
        end_date = date(year, 12, 31)
        period_label = str(year)
        prev_start_date = date(year-1, 1, 1)
        prev_end_date = date(year-1, 12, 31)

    meters = Meter.objects.filter(is_active=True)
    if group_id != 'all':
        meters = meters.filter(resource_type_id=group_id)

    table_data = []
    total_current = Decimal('0')
    total_prev = Decimal('0')

    for meter in meters:
        readings = meter.reading_set.filter(date__gte=start_date, date__lte=end_date)
        if not readings:
            continue
        if meter.is_multi_tariff:
            current = sum(r.total_consumption() for r in readings)
        else:
            current = readings.aggregate(total=Sum('consumption'))['total'] or Decimal('0')
        if current == 0:
            continue

        prev_readings = meter.reading_set.filter(date__gte=prev_start_date, date__lte=prev_end_date)
        if prev_readings:
            if meter.is_multi_tariff:
                prev = sum(r.total_consumption() for r in prev_readings)
            else:
                prev = prev_readings.aggregate(total=Sum('consumption'))['total'] or Decimal('0')
        else:
            prev = Decimal('0')

        if prev != 0:
            change = (current - prev) / prev * 100
            change_percent = change
            changetype = 'increase' if change > 0 else 'decrease' if change < 0 else 'stable'
        else:
            change_percent = 100 if current > 0 else 0
            changetype = 'increase' if current > 0 else 'stable'

        total_current += current
        total_prev += prev

        table_data.append({
            'meter': meter,
            'consumption': current,
            'prev_consumption': prev,
            'change_percent': change_percent,
            'changetype': changetype,
        })

    table_data.sort(key=lambda x: x['consumption'], reverse=True)

    if total_prev != 0:
        total_change = (total_current - total_prev) / total_prev * 100
        total_changetype = 'increase' if total_change > 0 else 'decrease' if total_change < 0 else 'stable'
    else:
        total_change = 100 if total_current > 0 else 0
        total_changetype = 'increase' if total_current > 0 else 'stable'

    units = {item['meter'].resource_type.unit for item in table_data}
    total_unit = units.pop() if len(units) == 1 else "ед."

    current_year = datetime.now().year
    years = range(current_year - 2, current_year + 1)
    resource_types = ResourceType.objects.all()

    # Логирование просмотра отчёта
    log_action(
        user=request.user,
        action='VIEW',
        model_name='Report',
        object_id='',
        details=f"Просмотр отчёта по потреблению за {period_label}, фильтр: {group_id}",
        request=request
    )

    context = {
        'table_data': table_data,
        'total_current': total_current,
        'total_prev': total_prev,
        'total_change': total_change,
        'total_changetype': total_changetype,
        'total_unit': total_unit,
        'period_type': period_type,
        'period_label': period_label,
        'selected_year': year,
        'selected_month': month,
        'selected_quarter': quarter,
        'selected_group': group_id,
        'years': years,
        'months': [(1, 'Январь'), (2, 'Февраль'), (3, 'Март'), (4, 'Апрель'),
                   (5, 'Май'), (6, 'Июнь'), (7, 'Июль'), (8, 'Август'),
                   (9, 'Сентябрь'), (10, 'Октябрь'), (11, 'Ноябрь'), (12, 'Декабрь')],
        'quarters': [1, 2, 3, 4],
        'resource_types': resource_types,
    }
    return render(request, 'energy/consumption_report.html', context)


# ------------------------------------------------------------
# Экспорт отчёта в Excel (с динамикой)
# ------------------------------------------------------------
@login_required
@viewer_required
def export_consumption_report(request):
    period_type = request.GET.get('period', 'month')
    year = int(request.GET.get('year', datetime.now().year))
    month = int(request.GET.get('month', datetime.now().month))
    quarter = int(request.GET.get('quarter', (datetime.now().month - 1) // 3 + 1))
    group_id = request.GET.get('group', 'all')

    month_names_ru = {
        1: 'Январь', 2: 'Февраль', 3: 'Март', 4: 'Апрель',
        5: 'Май', 6: 'Июнь', 7: 'Июль', 8: 'Август',
        9: 'Сентябрь', 10: 'Октябрь', 11: 'Ноябрь', 12: 'Декабрь'
    }

    if period_type == 'month':
        start_date = date(year, month, 1)
        end_date = start_date.replace(day=monthrange(year, month)[1])
        period_label = f"{month_names_ru[month]} {year}"
        prev_start_date = date(year-1, month, 1)
        prev_end_date = prev_start_date.replace(day=monthrange(year-1, month)[1])
    elif period_type == 'quarter':
        start_month = (quarter - 1) * 3 + 1
        start_date = date(year, start_month, 1)
        end_date = start_date + relativedelta(months=3) - timedelta(days=1)
        period_label = f"{quarter} квартал {year}"
        prev_start_date = date(year-1, start_month, 1)
        prev_end_date = prev_start_date + relativedelta(months=3) - timedelta(days=1)
    else:
        start_date = date(year, 1, 1)
        end_date = date(year, 12, 31)
        period_label = str(year)
        prev_start_date = date(year-1, 1, 1)
        prev_end_date = date(year-1, 12, 31)

    meters = Meter.objects.filter(is_active=True)
    if group_id != 'all':
        meters = meters.filter(resource_type_id=group_id)

    table_data = []
    total_current = Decimal('0')
    total_prev = Decimal('0')
    for meter in meters:
        readings = meter.reading_set.filter(date__gte=start_date, date__lte=end_date)
        if not readings:
            continue
        if meter.is_multi_tariff:
            current = sum(r.total_consumption() for r in readings)
        else:
            current = readings.aggregate(total=Sum('consumption'))['total'] or Decimal('0')
        if current == 0:
            continue

        prev_readings = meter.reading_set.filter(date__gte=prev_start_date, date__lte=prev_end_date)
        if prev_readings:
            if meter.is_multi_tariff:
                prev = sum(r.total_consumption() for r in prev_readings)
            else:
                prev = prev_readings.aggregate(total=Sum('consumption'))['total'] or Decimal('0')
        else:
            prev = Decimal('0')

        if prev != 0:
            change = (current - prev) / prev * 100
        else:
            change = 100 if current > 0 else 0

        total_current += current
        total_prev += prev
        table_data.append({
            'meter': meter,
            'consumption': current,
            'prev_consumption': prev,
            'change_percent': change,
        })

    table_data.sort(key=lambda x: x['consumption'], reverse=True)

    if total_prev != 0:
        total_change = (total_current - total_prev) / total_prev * 100
    else:
        total_change = 100 if total_current > 0 else 0

    # Логирование экспорта
    log_action(
        user=request.user,
        action='EXPORT',
        model_name='Report',
        object_id='',
        details=f"Экспорт отчёта по потреблению в Excel (период: {period_label}, фильтр: {group_id})",
        request=request
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Отчёт {period_label}"

    headers = [
        '№', 'Прибор учёта (серийный номер)', 'Тип ресурса', 'Тип тарифа',
        f'Потребление за {period_label} (ед.)',
        f'Потребление за аналогичный период прошлого года (ед.)',
        'Динамика, %'
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row, item in enumerate(table_data, 2):
        ws.cell(row=row, column=1, value=row-1)
        ws.cell(row=row, column=2, value=item['meter'].serial_number)
        ws.cell(row=row, column=3, value=f"{item['meter'].resource_type.name} ({item['meter'].resource_type.unit})")
        ws.cell(row=row, column=4, value="Многотарифный" if item['meter'].is_multi_tariff else "Одноставочный")
        ws.cell(row=row, column=5, value=float(item['consumption']))
        ws.cell(row=row, column=6, value=float(item['prev_consumption']))
        ws.cell(row=row, column=7, value=float(item['change_percent']))

    last_row = len(table_data) + 2
    ws.cell(row=last_row, column=4, value="ИТОГО").font = Font(bold=True)
    ws.cell(row=last_row, column=5, value=float(total_current)).font = Font(bold=True)
    ws.cell(row=last_row, column=6, value=float(total_prev)).font = Font(bold=True)
    ws.cell(row=last_row, column=7, value=float(total_change)).font = Font(bold=True)

    for col in range(1, 8):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 25

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="consumption_report_{period_label}.xlsx"'
    wb.save(response)
    return response


# ------------------------------------------------------------
# Отчёт по аномалиям потребления
# ------------------------------------------------------------
@login_required
@viewer_required
def anomaly_report(request):
    resource_type_id = request.GET.get('resource_type')
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    threshold = float(request.GET.get('threshold', 2.0))

    meters = Meter.objects.filter(is_active=True)
    if resource_type_id and resource_type_id != 'all':
        meters = meters.filter(resource_type_id=resource_type_id)

    anomalies = []
    for meter in meters:
        avg = get_avg_consumption(meter)
        if avg == 0:
            continue
        readings = meter.reading_set.order_by('-date')
        if start_date_str:
            readings = readings.filter(date__gte=start_date_str)
        if end_date_str:
            readings = readings.filter(date__lte=end_date_str)
        for reading in readings:
            consumption = reading.total_consumption()
            if is_anomaly(consumption, avg, threshold):
                anomalies.append({
                    'date': reading.date,
                    'meter': meter,
                    'consumption': consumption,
                    'avg': avg,
                    'ratio': float(consumption / avg) if avg else 0,
                })

    resource_types = ResourceType.objects.all()
    # Логирование просмотра отчёта
    log_action(
        user=request.user,
        action='VIEW',
        model_name='Report',
        object_id='',
        details=f"Просмотр отчёта по аномалиям (порог {threshold})",
        request=request
    )

    context = {
        'anomalies': anomalies,
        'resource_types': resource_types,
        'selected_resource_type': resource_type_id,
        'start_date': start_date_str,
        'end_date': end_date_str,
        'threshold': threshold,
    }
    return render(request, 'energy/anomaly_report.html', context)


# ------------------------------------------------------------
# Импорт показаний из Excel
# ------------------------------------------------------------
@login_required
@manager_required
def import_readings(request):
    if request.method == 'POST':
        form = ImportReadingsForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            dry_run = form.cleaned_data.get('dry_run', False)
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            ws = wb.active

            # Получаем заголовки (первая строка)
            headers = []
            for cell in ws[1]:
                if cell.value:
                    headers.append(str(cell.value).strip())
                else:
                    headers.append(None)
            headers = [h.strip() if h else '' for h in headers]

            # Детектируем колонки
            try:
                meter_col = headers.index('meter_serial') + 1
            except ValueError:
                meter_col = None
            try:
                date_col = headers.index('date') + 1
            except ValueError:
                date_col = None
            try:
                value_col = headers.index('value') + 1
            except ValueError:
                value_col = None

            zone_columns = []
            for i, h in enumerate(headers):
                if h and h not in ['meter_serial', 'date', 'value']:
                    if TariffComponent.objects.filter(name=h, is_multi_tariff_zone=True).exists():
                        zone_columns.append((i+1, h))
            if not meter_col or not date_col:
                messages.error(request, "Файл должен содержать колонки 'meter_serial' и 'date'")
                return redirect('energy:import_readings')
            if not value_col and not zone_columns:
                messages.error(request, "Файл должен содержать колонку 'value' для однотарифных или колонки с названиями зон для многотарифных счётчиков")
                return redirect('energy:import_readings')

            created_count = 0
            updated_count = 0
            errors = []

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row or all(cell is None for cell in row):
                    continue
                meter_serial = row[meter_col-1] if meter_col and len(row) >= meter_col else None
                date_str_val = row[date_col-1] if date_col and len(row) >= date_col else None
                if not meter_serial or not date_str_val:
                    errors.append(f"Строка {row_idx}: отсутствует серийный номер или дата")
                    continue
                try:
                    meter = Meter.objects.get(serial_number=str(meter_serial))
                except Meter.DoesNotExist:
                    errors.append(f"Строка {row_idx}: счётчик с серийным номером '{meter_serial}' не найден")
                    continue
                try:
                    if isinstance(date_str_val, datetime):
                        reading_date = date_str_val.date()
                    else:
                        reading_date = datetime.strptime(str(date_str_val), '%Y-%m-%d').date()
                except Exception:
                    errors.append(f"Строка {row_idx}: неверный формат даты (ожидается YYYY-MM-DD)")
                    continue

                if meter.is_multi_tariff:
                    if not zone_columns:
                        errors.append(f"Строка {row_idx}: счётчик многотарифный, но в файле нет колонок для зон")
                        continue
                    zone_values = {}
                    valid = True
                    for col_num, zone_name in zone_columns:
                        val = row[col_num-1] if len(row) >= col_num else None
                        if val is None:
                            errors.append(f"Строка {row_idx}: отсутствует значение для зоны '{zone_name}'")
                            valid = False
                            break
                        try:
                            zone_values[zone_name] = Decimal(str(val))
                        except Exception:
                            errors.append(f"Строка {row_idx}: неверное числовое значение для зоны '{zone_name}'")
                            valid = False
                            break
                    if not valid:
                        continue
                    components = {comp.name: comp for comp in TariffComponent.objects.filter(
                        resource_type=meter.resource_type, is_multi_tariff_zone=True,
                        valid_from__lte=reading_date
                    ).exclude(valid_to__lt=reading_date)}
                    missing_zones = set(zone_values.keys()) - set(components.keys())
                    if missing_zones:
                        errors.append(f"Строка {row_idx}: неизвестные зоны: {', '.join(missing_zones)}")
                        continue
                    if dry_run:
                        created_count += 1
                        continue
                    reading, created = Reading.objects.update_or_create(
                        meter=meter, date=reading_date,
                        defaults={'value': None}
                    )
                    for zone_name, val in zone_values.items():
                        comp = components[zone_name]
                        ZoneReading.objects.update_or_create(
                            reading=reading, tariff_component=comp,
                            defaults={'value': val}
                        )
                    if not dry_run:
                        meter.recalc_consumption()
                    if created:
                        created_count += 1
                    else:
                        updated_count += 1
                    # Логирование для каждой успешной строки (можно сгруппировать, но для простоты оставим)
                    log_action(
                        user=request.user,
                        action='IMPORT',
                        model_name='Reading',
                        object_id=reading.pk,
                        details=f"Импорт показаний для {meter.serial_number} от {reading_date}",
                        request=request
                    )
                else:
                    if not value_col:
                        errors.append(f"Строка {row_idx}: для однотарифного счётчика нужна колонка 'value'")
                        continue
                    val = row[value_col-1] if len(row) >= value_col else None
                    if val is None:
                        errors.append(f"Строка {row_idx}: отсутствует значение показания")
                        continue
                    try:
                        value = Decimal(str(val))
                    except Exception:
                        errors.append(f"Строка {row_idx}: неверное числовое значение для показания")
                        continue
                    if dry_run:
                        created_count += 1
                        continue
                    reading, created = Reading.objects.update_or_create(
                        meter=meter, date=reading_date,
                        defaults={'value': value}
                    )
                    if not dry_run:
                        meter.recalc_consumption()
                    if created:
                        created_count += 1
                    else:
                        updated_count += 1
                    log_action(
                        user=request.user,
                        action='IMPORT',
                        model_name='Reading',
                        object_id=reading.pk,
                        details=f"Импорт показаний для {meter.serial_number} от {reading_date}",
                        request=request
                    )

            if errors:
                messages.warning(request, f"Импорт завершён с ошибками. Создано/обновлено: {created_count}, пропущено строк: {len(errors)}. Ошибки: {', '.join(errors[:10])}")
            else:
                messages.success(request, f"Импорт успешно завершён. Создано: {created_count}, обновлено: {updated_count}")
            if dry_run:
                messages.info(request, "Режим проверки (dry run) – данные не сохранены")
            return redirect('energy:import_readings')
    else:
        form = ImportReadingsForm()
    return render(request, 'energy/import_readings.html', {'form': form})


# ------------------------------------------------------------
# Скачивание шаблона для импорта
# ------------------------------------------------------------
@login_required
@manager_required
def download_import_template(request):
    # Логирование скачивания шаблона
    log_action(
        user=request.user,
        action='EXPORT',
        model_name='Template',
        object_id='',
        details="Скачивание шаблона импорта показаний",
        request=request
    )
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Импорт показаний"

    headers = ['meter_serial', 'date', 'value']
    example_zones = TariffComponent.objects.filter(is_multi_tariff_zone=True).values_list('name', flat=True)[:2]
    for zone in example_zones:
        headers.append(zone)
    ws.append(headers)

    meter = Meter.objects.filter(is_active=True).first()
    if meter:
        ws.append([meter.serial_number, '2026-01-15', '123.45'] + [''] * len(example_zones))
        ws.append([meter.serial_number, '2026-02-15', '234.56'] + [''] * len(example_zones))
    else:
        ws.append(['СЧ-001', '2026-01-15', '123.45'] + [''] * len(example_zones))
        ws.append(['СЧ-001', '2026-02-15', '234.56'] + [''] * len(example_zones))

    for col in range(1, len(headers)+1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="import_readings_template.xlsx"'
    wb.save(response)
    return response

@login_required
@viewer_required
def export_readings(request):
    meter_id = request.GET.get('meter')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    if not meter_id:
        meters = Meter.objects.all()
        if not can_view_all_meters(request.user):
            meters = meters.filter(user=request.user)
        return render(request, 'energy/export_readings.html', {'meters': meters})

    meter = get_object_or_404(Meter, pk=meter_id)
    if not can_view_meter(request.user, meter):
        messages.error(request, "Нет доступа к счётчику")
        return redirect('energy:meter_list')

    readings = meter.reading_set.all().order_by('-date')
    if start_date:
        readings = readings.filter(date__gte=start_date)
    if end_date:
        readings = readings.filter(date__lte=end_date)

    if not readings.exists():
        messages.warning(request, "Нет показаний за выбранный период")
        return redirect('energy:export_readings')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "История показаний"

    if meter.is_multi_tariff:
        zones = TariffComponent.objects.filter(resource_type=meter.resource_type, is_multi_tariff_zone=True).order_by('name')
        headers = ['Дата'] + [zone.name for zone in zones] + ['Суммарное потребление (ед.)']
        ws.append(headers)
        for reading in readings:
            row = [reading.date.strftime('%d.%m.%Y')]
            zone_values = {zr.tariff_component.name: zr.value for zr in reading.zone_readings.all()}
            for zone in zones:
                row.append(zone_values.get(zone.name, ''))
            row.append(float(reading.total_consumption()))
            ws.append(row)
    else:
        headers = ['Дата', 'Показание (суммарное)', 'Потребление (ед.)']
        ws.append(headers)
        for reading in readings:
            ws.append([
                reading.date.strftime('%d.%m.%Y'),
                float(reading.value) if reading.value else '',
                float(reading.consumption) if reading.consumption else ''
            ])

    for col in range(1, len(headers)+1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"readings_{meter.serial_number}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response